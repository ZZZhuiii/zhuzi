"""
=======================================================================
常用库速览 —— 每个库最核心的用法（逐行注释版）
初次接触不用全懂，跑一遍看效果，用到哪个回来看哪个
=======================================================================
"""

# 一行搞定路径：导入后直接用 ROOT / OUT / DATA
from _env import *


# ============================================================
# 1. pandas —— 数据处理之王
# ============================================================
# pd 是 pandas 的约定俗成简称，行业里大家都这么写
import pandas as pd

# DataFrame 是 pandas 最核心的东西：一张"表格"
# 可以理解为一个超级 Excel 工作表，能装几百万行也不卡
df = pd.DataFrame({

    # 字典的每个 key 就是一列，每个 value 的列表就是这一列的值
    "商品": ["A", "B", "C"],     # 第一列：商品名（文本）
    "销量": [100, 200, 150],     # 第二列：销量（数字）
    "单价": [10, 20, 15],        # 第三列：单价（数字）
})

# 直接像 Excel 公式一样，整列相乘，生成新的一列
# 销量 × 单价 = 销售额
df["销售额"] = df["销量"] * df["单价"]

print("pandas 表格计算：")
print(df)    # 直接打印，自动对齐排版
print()


# ============================================================
# 2. openpyxl —— 读写 Excel 文件
# ============================================================
from openpyxl import Workbook

wb = Workbook()          # 创建一个新的 Excel 工作簿（相当于新建一个 .xlsx 文件）
ws = wb.active           # 获取默认激活的那张工作表（Sheet）

ws["A1"] = "Hello"       # A1 单元格写入 "Hello"
ws["B1"] = "Python"      # B1 单元格写入 "Python"

# 保存到 output/hello.xlsx
# OUT / "hello.xlsx" 就是 python/output/hello.xlsx
wb.save(OUT / "hello.xlsx")
print("openpyxl 已生成 hello.xlsx  →  output/")


# ============================================================
# 3. requests —— 发送网络请求（浏览器能看的，它就能拿）
# ============================================================
import requests

# 向百度首页发送 GET 请求（相当于在浏览器地址栏输入网址按回车）
r = requests.get("https://www.baidu.com")

# status_code 是 HTTP 状态码
# 200 = 成功  404 = 找不到  500 = 服务器报错
print("百度首页状态码：", r.status_code)


# ============================================================
# 4. BeautifulSoup —— 解析网页 HTML，提取你要的信息
# ============================================================
from bs4 import BeautifulSoup

# 一段简单的 HTML 字符串（实际使用时通常是 requests 获取的网页内容）
html = "<html><body><h1>标题</h1><p>段落</p></body></html>"

# 用 html.parser 解析这段 HTML
# soup 现在是一个可以查询的对象
soup = BeautifulSoup(html, "html.parser")

# .h1 取出第一个 <h1> 标签，.text 取出标签里的文字内容
print("解析到标题：", soup.h1.text)


# ============================================================
# 5. Pillow —— 图片处理（改尺寸、加水印、换格式）
# ============================================================
from PIL import Image, ImageDraw

# 新建一张 200×100 的纯蓝色图片（RGB 色彩模式）
img = Image.new("RGB", (200, 100), "blue")

# 创建一个"画笔"，可以在图片上写字、画线
draw = ImageDraw.Draw(img)

# 在坐标 (10, 40) 处写文字 "Python!"，白色
draw.text((10, 40), "Python!", fill="white")

# 保存到 output/demo.png
img.save(OUT / "demo.png")
print("Pillow 已生成 demo.png  →  output/")
